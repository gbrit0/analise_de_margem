from setup import settings
from .filters import NotaFilter
from .models import (
    Justificativa, 
    Nota, 
    Custo, 
    Margem, 
    Nf_Has_Justificativa,
    Log_Comentario,
    OP,
    Custo2_OP,
    MesBloqueado,
    LogBloqueioMes,
    PreferenciaGlobalColunasNota
)

from users.models import CustomUser

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

import os
import json
import locale
import pyodbc
import pymysql
import datetime
from decimal import Decimal
from typing import Any, TypedDict
from dbutils.pooled_db import PooledDB
from dateutil.relativedelta import relativedelta

from django.utils import timezone
from django.shortcuts import render
from django.views.generic import ListView
from django_filters.views import FilterView
from django.utils.dateparse import parse_date
from django.db.models import Sum, Count, Q, Avg
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.views.decorators.cache import cache_page
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models.functions import Coalesce, ExtractYear, ExtractMonth
from django.views.decorators.http import require_http_methods, require_POST
from django.db.models import OuterRef, Subquery, DecimalField, Case, When, F, Value, ExpressionWrapper

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

NOTA_GRID_COLUMNS = [
    {'key': 'filial', 'label': 'Filial', 'index': 1, 'default_visible': True},
    {'key': 'nota', 'label': 'Nota', 'index': 2, 'default_visible': True},
    {'key': 'emissao', 'label': 'Emissão', 'index': 3, 'default_visible': True},
    {'key': 'pedido', 'label': 'Nº Pedido', 'index': 4, 'default_visible': True},
    {'key': 'vendedor', 'label': 'Vendedor', 'index': 5, 'default_visible': True},
    {'key': 'cliente', 'label': 'Cliente', 'index': 6, 'default_visible': True},
    {'key': 'cfop', 'label': 'CFOP', 'index': 7, 'default_visible': False},
    {'key': 'tipo', 'label': 'Tipo', 'index': 8, 'default_visible': False},
    {'key': 'produto', 'label': 'Produto', 'index': 9, 'default_visible': True},
    {'key': 'lote', 'label': 'Lote', 'index': 10, 'default_visible': False},
    {'key': 'quantidade', 'label': 'Qtd', 'index': 11, 'default_visible': False},
    {'key': 'valor_contabil', 'label': 'Vl. Contábil', 'index': 12, 'default_visible': True},
    {'key': 'custo', 'label': 'Custo', 'index': 13, 'default_visible': True},
    {'key': 'tabela_preco', 'label': 'Tabela Preço', 'index': 14, 'default_visible': False},
    {'key': 'preco_tabela', 'label': 'Preço Tabela', 'index': 15, 'default_visible': False},
    {'key': 'margem_bruta', 'label': 'Margem Bruta', 'index': 16, 'default_visible': True},
    {'key': 'margem_percentual', 'label': 'Margem %', 'index': 17, 'default_visible': True},
    {'key': 'estoque', 'label': 'Estoque?', 'index': 18, 'default_visible': False},
    {'key': 'duplicata', 'label': 'Duplicata?', 'index': 19, 'default_visible': False},
    {'key': 'armazem', 'label': 'Armazém', 'index': 20, 'default_visible': False},
    {'key': 'grp_amar_ctb', 'label': 'Gp. Amar. CTB', 'index': 21, 'default_visible': False},
    {'key': 'uf', 'label': 'UF', 'index': 22, 'default_visible': False},
    {'key': 'valor_unitario', 'label': 'Vl. Unit', 'index': 23, 'default_visible': False},
    {'key': 'ipi', 'label': 'IPI', 'index': 24, 'default_visible': False},
    {'key': 'imp5', 'label': 'Imp5', 'index': 25, 'default_visible': False},
    {'key': 'imp6', 'label': 'Imp6', 'index': 26, 'default_visible': False},
    {'key': 'difal', 'label': 'Difal', 'index': 27, 'default_visible': False},
    {'key': 'icms', 'label': 'ICMS', 'index': 28, 'default_visible': False},
    {'key': 'aliq_icms', 'label': 'Aliq %', 'index': 29, 'default_visible': False},
    {'key': 'justificativa', 'label': 'Justificativa', 'index': 30, 'default_visible': True},
    {'key': 'comentario', 'label': 'Comentário', 'index': 31, 'default_visible': False},
]

NOTA_GRID_COLUMN_KEYS = {column['key'] for column in NOTA_GRID_COLUMNS}
NOTA_GRID_MAIN_COLUMN_KEYS = {
    'filial',
    'nota',
    'pedido',
    'vendedor',
    'cliente',
    'produto',
    'valor_contabil',
    'custo',
    'margem_bruta',
    'margem_percentual',
    'justificativa',
}
NOTA_GRID_DEFAULT_VISIBLE = [
    column['key'] for column in NOTA_GRID_COLUMNS if column['default_visible']
]


def get_nota_grid_visible_columns():
    preferencia = PreferenciaGlobalColunasNota.objects.first()
    if not preferencia or not isinstance(preferencia.colunas_visiveis, list):
        return NOTA_GRID_DEFAULT_VISIBLE

    colunas = [
        coluna for coluna in preferencia.colunas_visiveis
        if coluna in NOTA_GRID_COLUMN_KEYS
    ]
    return colunas or NOTA_GRID_DEFAULT_VISIBLE

pool = PooledDB(
    creator=pyodbc,
    maxconnections=10, # Como vamos usar lotes, não precisamos de tantas conexões simultâneas
    mincached=2,
    blocking=True,
    driver='{ODBC Driver 17 for SQL Server}',
    server=f'{os.getenv("PROTHEUS_DB_HOST")}',
    database=f'{os.getenv("PROTHEUS_DB_DATABASE")}',
    uid=f'{os.getenv("PROTHEUS_DB_USER")}',
    pwd=f'{os.getenv("PROTHEUS_DB_PASSWORD")}'
)

pool_mysql = PooledDB(
    creator=pymysql,
    maxconnections=10, # Como vamos usar lotes, não precisamos de tantas conexões simultâneas
    mincached=2,
    blocking=True,
    host=f'{os.getenv("DB_HOST")}',
    port=int(os.getenv("DB_PORT")),
    user=f'{os.getenv("DB_USER")}',
    password=f'{os.getenv("DB_PASSWORD")}',
    database=f'{os.getenv("DB_NAME")}'
)

class NotasListView(LoginRequiredMixin, FilterView, ListView):
    mes_anterior = datetime.date.today() - relativedelta(month=1)
    mes_anterior = mes_anterior.strftime("%Y-%m")
    model = Nota
    # paginate_by = 25
    filterset_class = NotaFilter
    ordering = ['-data_emissao']
    login_url = 'login/' 
    context_object_name = 'nota_list'
    template_name = 'notas/nota_list.html'
    redirect_field_name = f'/?data_emissao_month={mes_anterior}' # Filtrar as notas do mes anterior depois do login
        
    def get_queryset(self):
        cfops_especiais = ['5101', '6101', '5116', '6116', '6107']
        # 1. Definimos a subquery para buscar o Custo relacionado
        # Filtramos onde a chave do Custo é igual à chave da Nota atual (OuterRef('pk'))
        custo_mais_recente = Custo.objects.filter(
            chave=OuterRef('pk')
        ).order_by('-data_cadastro').values('valor')[:1]
        
        ultimo_custo = Custo.objects.filter(
            chave=OuterRef('pk')
        ).order_by('data_cadastro').values('valor')[:1]
        
        justificativa_mais_recente = Nf_Has_Justificativa.objects.filter(
            nf=OuterRef('pk')
        ).order_by('-data_cadastro').values('justificativa')[:1]
        
        # 2. Anotamos o queryset principal da Nota com esse valor
        queryset = Nota.objects.filter(delete=False).annotate(
            custo_mais_recente=Coalesce(
                Subquery(custo_mais_recente), 
                0, 
                output_field=DecimalField()
            ),
            ultimo_custo_valor=Coalesce(
                Subquery(ultimo_custo), 
                0, 
                output_field=DecimalField()
            ),
            icms_calculado=Case(
                When(cfop__in=cfops_especiais, then=F('base_icms') * 0.047), # 4,7% do ICMS
                default=F('valor_icms'),                                     # ICMS cheio
                output_field=DecimalField()
            ),
            margem_bruta=ExpressionWrapper(
                F('valor_contabil') 
                - F('custo_mais_recente') 
                - F('valor_ipi') 
                - F('valor_imp5') 
                - F('valor_imp6') 
                - F('valor_icms_difal') 
                - F('icms_calculado'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            ),
            margem_percentual=ExpressionWrapper(
                F('margem_bruta') 
                / F('valor_contabil') * 100,
                output_field=DecimalField(max_digits=4, decimal_places=2)
            ),
            justificativa=justificativa_mais_recente
        )
        
        filiais_selecionadas = self.request.GET.getlist('filial')
        if filiais_selecionadas:
            queryset = queryset.filter(filial__in=filiais_selecionadas)
            
        justificativas_selecionadas = self.request.GET.getlist('justificativa_filtro')
        if justificativas_selecionadas:
            queries = Q()
            ids = []
            for j in justificativas_selecionadas:
                if j == 'None':
                    queries |= Q(justificativa__isnull=True)
                else:
                    ids.append(j)
            
            if ids:
                queries |= Q(justificativa__in=ids)
            
            queryset = queryset.filter(queries)
            
        return queryset.order_by('-data_emissao')
    
    def get_context_data(self, **kwargs):
        # Chama a implementação base primeiro para pegar o contexto padrão (incluindo o queryset)
        context = super().get_context_data(**kwargs)
        
        # Adicionamos a lista de justificativas ativas para popular os selects no HTML
        context['lista_justificativas'] = Justificativa.objects.filter(ativo=True)
        context['nota_grid_columns'] = NOTA_GRID_COLUMNS
        context['nota_grid_visible_columns'] = get_nota_grid_visible_columns()
        context['nota_grid_main_column_keys'] = list(NOTA_GRID_MAIN_COLUMN_KEYS)
        
        # Monta lista de meses/anos disponíveis (format YYYY-MM)
        meses_qs = Nota.objects.filter(delete=False).annotate(
            year=ExtractYear('data_emissao'),
            month=ExtractMonth('data_emissao')
        ).values('year', 'month').distinct().order_by('-year', '-month')

        meses = []
        for m in meses_qs:
            y = int(m['year'])
            mo = int(m['month'])
            meses.append({
                'value': f"{y}-{mo:02d}",
                'label': f"{mo:02d}/{y}"
            })
            
        context['meses_disponiveis'] = meses
        # Valor selecionado atualmente (vindo dos GET params)
        selected = self.request.GET.get('data_emissao_month', '')

        filiais_selecionadas = self.request.GET.getlist('filial')
        context['filiais_selecionadas'] = filiais_selecionadas
        
        justificativas_selecionadas = self.request.GET.getlist('justificativa_filtro')
        context['justificativas_selecionadas'] = justificativas_selecionadas

        # Se não veio valor no GET, tenta preencher com mês atual (se houver dados),
        # caso contrário preenche com o mês mais recente disponível.
        if not selected:
            hoje = datetime.datetime.today()
            atual = f"{hoje.year}-{hoje.month:02d}"
            # se atual estiver na lista de meses disponiveis, usa; senão usa o primeiro disponível
            valores = [m['value'] for m in meses]
            if atual in valores:
                selected = atual
            elif len(valores) > 0:
                selected = valores[0]

        context['selected_month'] = selected
        
        # context['mes_anterior'] = (datetime.date.today() - relativedelta(month=1)).strftime("%Y-%m")
        filiais = Nota.objects.filter(delete=False).values('filial', 'nome_filial').distinct()
        
        context['filiais'] = filiais
        
        with pool_mysql.connection() as con:
            with con.cursor() as cursor:
                query = """SELECT cod_cliente, loja_cliente FROM analise_margem.cliente_parceiro;"""
                cursor.execute(query)
                clientes_parceiros_raw = cursor.fetchall()
        # 4459286002
        # 1. Cria um SET de tuplas para busca extremamente rápida.
        # Assumindo que o fetchall() retorna tuplas onde [0] é cod_cliente e [1] é loja.
        # (Se seu cursor retornar dicionários, use: c['cod_cliente'], c['loja'])
        set_parceiros = {(c[0], c[1]) for c in clientes_parceiros_raw}
        
        set_intercompany = {
            ('44592860', '0001'), # Agrogera BA
            ('04675878', '0001'), # BRG Matriz
            ('44592860', '0002'), # Agrogera BA Filial GO
            ('27379581', '0004'), # GRID MS - INOCENCIA
            ('27379581', '0001'), # GRID GO
            ('27379581', '0002'), # GRID MG
            ('27379581', '0003'), # GRID PA
        }
        
        # 2. Verifica a lista de notas atual (object_list do ListView)
        # Se você usa um context_object_name diferente, troque 'object_list' pelo seu nome
        for nota in context['object_list']:
            # Cria um atributo dinâmico "is_parceiro" na nota
            # Se a tupla (cod_cliente, loja) da nota existir no SET, recebe True.
            if (nota.cod_cliente, nota.loja) in set_intercompany:
                nota.is_intercompany = True
                justificativa_nf = Nf_Has_Justificativa.objects.filter(nf=nota).last()
                if not justificativa_nf:
                    Nf_Has_Justificativa.objects.create(
                        nf=nota,
                        justificativa=Justificativa.objects.get(texto='Venda Intercompany'),
                        usuario=CustomUser.objects.get(id=2)
                    )
            elif (nota.cod_cliente, nota.loja) in set_parceiros:
                nota.is_parceiro = True
                margem = Margem.objects.filter(chave=nota).last()
                
                if margem.margem_bruta_percentual < 0.27 :
                    if margem.margem_bruta_percentual > 0.15:
                        justificativa_nf = Nf_Has_Justificativa.objects.filter(nf=nota).last()
                        if not justificativa_nf:
                            Nf_Has_Justificativa.objects.create(
                                nf=nota,
                                justificativa=Justificativa.objects.get(texto='OK. Margem Parceiro.'),
                                usuario=CustomUser.objects.get(id=2)
                            )
                        
            else:
                nota.is_parceiro = False
                nota.is_intercompany = False

        return context
    
@login_required
@require_POST
def atualizar_custo_api(request):
    try:
        data = json.loads(request.body)
        nota_chave = data.get('chave')
        novo_valor_custo = data.get('valor')
        
        # 1. Validação básica
        if not nota_chave or novo_valor_custo is None:
            return JsonResponse({'error': 'Dados inválidos'}, status=400)

        valor_custo_decimal = Decimal(str(novo_valor_custo).replace(',', '.'))
        
        # 2. Busca a Nota
        nota = get_object_or_404(Nota, pk=nota_chave)

        # Validação de Bloqueio do Mês
        mes_ano_nota = nota.data_emissao.strftime("%Y-%m")
        mes_status = MesBloqueado.objects.filter(mes_ano=mes_ano_nota).first()
        if mes_status and mes_status.bloqueado:
            return JsonResponse({'error': f'Mês {mes_ano_nota} fechado para edição.'}, status=403)
        
        # 3. Verifica regra de bloqueio: se a nota pertence a mês anterior e hoje é > 03, impede atualização
        nota_month = nota.data_emissao.month
        nota_year = nota.data_emissao.year
        today = datetime.date.today()
        # Se a data da nota está em mês/ano anterior ao atual e hoje é após o dia 3, bloqueia
        if (today.year, today.month) > (nota_year, nota_month) and today.day > 3:
            return JsonResponse({'error': 'Atualização de custo bloqueada após fechamento do dia 03'}, status=403)
        
        if Custo.objects.filter(chave=nota).exists():
            if not request.user.is_superuser:
                return JsonResponse({'success': False, 'error': 'Custo já cadastrado. Somente administradores podem alterar.'}, status=403)

        # 3. Cria um NOVO registro de custo (preservando histórico)
        custo = Custo.objects.create(
            chave=nota,
            valor=valor_custo_decimal,
            usuario=request.user if request.user else 2  # usuário system
        )
               
        # Definição do ICMS baseada no CFOP
        cfops_especiais = ['5101', '6101', '5116', '6116', '6107']
        if nota.cfop in cfops_especiais:
            val_icms = nota.base_icms * Decimal('0.047')
        else:
            val_icms = nota.valor_icms

        # Cálculo: Valor Contábil - Custo Novo - Impostos
        margem_bruta = (
            nota.valor_contabil 
            - valor_custo_decimal 
            - nota.valor_ipi 
            - nota.valor_imp5 
            - nota.valor_imp6 
            - nota.valor_icms_difal 
            - val_icms
        )

        # Cálculo da Margem Percentual (evitando divisão por zero)
        margem_percentual = 0
        if nota.valor_contabil > 0:
            margem_percentual = (margem_bruta / nota.valor_contabil)

        margem = Margem.objects.create(
            chave=nota,
            custo=custo,
            margem_bruta=margem_bruta,
            margem_bruta_percentual=margem_percentual
        )
        
        # 5. Retorna os dados formatados para o Frontend
        return JsonResponse({
            'success': True,
            'margem_bruta': f'{locale.currency(margem_bruta, grouping=True)}',
            'margem_percentual': f'{margem_percentual*100:.2f}%'
        })

    except Exception as e:
        raise e
        # return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_POST
def atualizar_custo2_op_api(request):
    try:
        data = json.loads(request.body)
        id_op = data.get('id_op')
        novo_valor = data.get('valor')
        
        if not id_op or novo_valor is None:
            return JsonResponse({'error': 'Dados inválidos'}, status=400)

        valor_decimal = Decimal(str(novo_valor).replace(',', '.'))
        
        op_obj = OP.objects.filter(id_op=id_op).first()
        if not op_obj:
            return JsonResponse({'error': 'OP não encontrada'}, status=404)
        
        Custo2_OP.objects.create(
            op=op_obj,
            valor=valor_decimal,
            usuario=request.user if request.user else CustomUser.objects.get(id=2)
        )
        
        # O histórico também salva na tabela principal OP para ser fácil resgatar depois
        op_obj.custo_2 = valor_decimal
        op_obj.save(update_fields=['custo_2'])
        
        return JsonResponse({
            'success': True,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@login_required
@require_POST
def atualizar_comentario_api(request):
    try:
        data = json.loads(request.body)
        nota_chave = data.get('chave')
        novo_comentario = data.get('comentario')

        nota = get_object_or_404(Nota, pk=nota_chave)
        nota.comentario = novo_comentario
        nota.save(update_fields=['comentario'])

        Log_Comentario.objects.create(
            nf=nota,
            comentario=novo_comentario,
            usuario=request.user
        )

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def atualizar_justificativa_api(request):
    try:
        data = json.loads(request.body)
        nota_chave = data.get('chave')
        nova_justificativa_id = data.get('justificativa')

        nota = get_object_or_404(Nota, pk=nota_chave)
        
        # Validação de Bloqueio do Mês
        mes_ano_nota = nota.data_emissao.strftime("%Y-%m")
        mes_status = MesBloqueado.objects.filter(mes_ano=mes_ano_nota).first()
        if mes_status and mes_status.bloqueado:
            return JsonResponse({'success': False, 'error': f'Mês {mes_ano_nota} fechado para edição.'}, status=403)

        if Nf_Has_Justificativa.objects.filter(nf=nota).exists():
            if not request.user.is_superuser:
                return JsonResponse({'success': False, 'error': 'Justificativa já cadastrada. Somente administradores podem alterar.'}, status=403)

        # 1. Buscamos a justificativa que o usuário clicou
        justificativa = get_object_or_404(Justificativa, id=nova_justificativa_id)

        # 2. VERIFICAÇÃO: Se o texto for 'Limpar / Sem Justificativa', 
        # nós apenas deletamos e retornamos.
        # if 'Limpar' in justificativa.texto:
        #     Nf_Has_Justificativa.objects.filter(nf=nota).delete()
        #     return JsonResponse({'success': True, 'action': 'cleared'})

        # 3. Se NÃO for limpar, primeiro limpamos o que tinha antes (para não duplicar)
        # e depois criamos a nova relação
        # Nf_Has_Justificativa.objects.filter(nf=nota).delete()
        
        Nf_Has_Justificativa.objects.create(
            usuario=request.user,
            nf=nota,
            justificativa=justificativa
        )

        return JsonResponse({'success': True, 'justificativa_id': justificativa.id})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def dashboard_view(request):
    selected = request.GET.get('data_emissao_month', '')
    
    meses_qs = Nota.objects.filter(delete=False).annotate(
        year=ExtractYear('data_emissao'),
        month=ExtractMonth('data_emissao')
    ).values('year', 'month').distinct().order_by('-year', '-month')

    meses = []
    for m in meses_qs:
        y = int(m['year'])
        mo = int(m['month'])
        meses.append({
            'value': f"{y}-{mo:02d}",
            'label': f"{mo:02d}/{y}"
        })

    # Se não veio valor no GET, tenta preencher com mês atual (se houver dados),
    # caso contrário preenche com o mês mais recente disponível.
    if not selected:
        hoje = datetime.datetime.today()
        atual = f"{hoje.year}-{hoje.month:02d}"
        valores = [m['value'] for m in meses]
        if atual in valores:
            selected = atual
        elif len(valores) > 0:
            selected = valores[0]
        
    filiais = Nota.objects.filter(delete=False).values('filial', 'nome_filial').distinct()
    
    context = {
        'filiais': filiais,
        'meses_disponiveis': meses,
        'selected_month': selected
    }
    
    return render(request, 'notas/estatisticas.html', context)

@login_required
def dados_vendas_api(request):
    
    meses_str = request.GET.get('meses')
    filiais_str = request.GET.get('filiais')
        
    queryset = Nota.objects.filter(delete=False)
    
    if meses_str:
        meses_list = meses_str.split(',')
        queries = Q()
        for mes in meses_list:
            if '-' in mes:
                try:
                    y, m = mes.split('-')
                    queries |= Q(data_emissao__year=int(y), data_emissao__month=int(m))
                except ValueError:
                    pass
        if queries:
            queryset = queryset.filter(queries)
            
    if filiais_str:
        filiais_list = filiais_str.split(',')
        queryset = queryset.filter(filial__in=filiais_list)
    
    subquery_margem_percentual = Margem.objects.filter(
        chave=OuterRef('chave')
    ).values('margem_bruta_percentual').order_by('-custo__data_cadastro')[:1] # O Gemini disse: Para ordenar pelo atributo cadastro da tabela Custo (que possui a Foreign Key para Margem), você precisa utilizar a sintaxe de "follow relationship" do Django, que utiliza o duplo sublinhado (__).
    
    subquery_custo = Custo.objects.filter(
            chave=OuterRef('pk')
        ).order_by('-data_cadastro').values('valor')[:1]
    
    subquery_margem = Margem.objects.filter(
        chave=OuterRef('chave')
    ).values('margem_bruta').order_by('-custo__data_cadastro')[:1]

    subquery_justificaticas = Nf_Has_Justificativa.objects.filter(
        nf=OuterRef('pk')
    ).values('justificativa')[:1]
    
    subquery_justificativas_texto = Nf_Has_Justificativa.objects.filter(
        nf=OuterRef('pk')
    ).order_by('-data_cadastro').values('justificativa__texto')[:1]
    
    with pool_mysql.connection() as con:
        with con.cursor() as cursor:
            query = """SELECT cod_cliente, loja_cliente FROM analise_margem.cliente_parceiro;"""
            cursor.execute(query)
            clientes_parceiros_raw = cursor.fetchall()
            
    set_parceiros = {(c[0], c[1]) for c in clientes_parceiros_raw}

    notas_stats = queryset.annotate(
        margem_pct=Subquery(subquery_margem_percentual),
        just_texto=Subquery(subquery_justificativas_texto)
    ).values('cod_cliente', 'loja', 'valor_contabil', 'margem_pct', 'just_texto')
    
    total_vendas_periodo = Decimal('0.0')
    stats_just = {}
    total_notas_abaixo_margem_global = 0

    for nota in notas_stats:
        val_contabil = nota['valor_contabil'] or Decimal('0.0')
        total_vendas_periodo += val_contabil
        
        is_parceiro = (nota['cod_cliente'], nota['loja']) in set_parceiros
        limite_margem = 0.15 if is_parceiro else 0.27
        margem_atual = nota['margem_pct']
        
        abaixo_margem = False
        if margem_atual is not None and float(margem_atual) < limite_margem:
            abaixo_margem = True
            total_notas_abaixo_margem_global += 1
            
        j_texto = nota['just_texto']
        if j_texto:
            if j_texto not in stats_just:
                stats_just[j_texto] = {
                    'contagem': 0,
                    'vendas_total': Decimal('0.0'),
                    'abaixo_margem': 0
                }
            stats_just[j_texto]['contagem'] += 1
            stats_just[j_texto]['vendas_total'] += val_contabil
            if abaixo_margem:
                stats_just[j_texto]['abaixo_margem'] += 1

    estatisticas_justificativas = []
    total_vendas_float = float(total_vendas_periodo)
    
    for j_texto, data in stats_just.items():
        repres_vendas = (float(data['vendas_total']) / total_vendas_float * 100) if total_vendas_float > 0 else 0
        perc_abaixo = (data['abaixo_margem'] / total_notas_abaixo_margem_global * 100) if total_notas_abaixo_margem_global > 0 else 0
        
        estatisticas_justificativas.append({
            'justificativa': j_texto,
            'contagem': data['contagem'],
            'representatividade_vendas': round(repres_vendas, 2),
            'percentual_abaixo_margem': round(perc_abaixo, 2)
        })
        
    estatisticas_justificativas.sort(key=lambda x: x['contagem'], reverse=True)

    estatisticas_vendedores_qs = queryset.values('vendedor').annotate(
        total_vendas=Sum('valor_contabil'),
        margem_total=Sum(subquery_margem),
        custo_total=Sum(subquery_custo),
        qtd_notas=Count('chave')
    ).order_by('-total_vendas')

    estatisticas_vendedores = []
    for item in estatisticas_vendedores_qs:
        total_vendedor = item['total_vendas'] or Decimal('0.0')
        margem_vendedor = item['margem_total'] or Decimal('0.0')
        custo_vendedor = item['custo_total'] or Decimal('0.0')
        margem_percentual_vendedor = 0
        if total_vendedor:
            margem_percentual_vendedor = (margem_vendedor / total_vendedor) * 100

        estatisticas_vendedores.append({
            'vendedor': item['vendedor'] or 'Sem vendedor',
            'qtd_notas': item['qtd_notas'],
            'total_vendas': float(total_vendedor),
            'custo_total': float(custo_vendedor),
            'margem_total': float(margem_vendedor),
            'margem_percentual': round(float(margem_percentual_vendedor), 2)
        })

    top_vendedores = [item['vendedor'] for item in estatisticas_vendedores_qs[:5] if item['vendedor']]
    estatisticas_vendedores_periodo = []
    if top_vendedores:
        estatisticas_vendedores_periodo_qs = queryset.filter(
            vendedor__in=top_vendedores
        ).annotate(
            mes=TruncMonth('data_emissao')
        ).values('mes', 'vendedor').annotate(
            total_vendas=Sum('valor_contabil'),
            margem_total=Sum(subquery_margem)
        ).order_by('mes', 'vendedor')

        for item in estatisticas_vendedores_periodo_qs:
            total_vendedor_periodo = item['total_vendas'] or Decimal('0.0')
            margem_vendedor_periodo = item['margem_total'] or Decimal('0.0')
            margem_percentual_periodo = 0
            if total_vendedor_periodo:
                margem_percentual_periodo = (margem_vendedor_periodo / total_vendedor_periodo) * 100

            estatisticas_vendedores_periodo.append({
                'mes': item['mes'].strftime('%b/%Y'),
                'vendedor': item['vendedor'] or 'Sem vendedor',
                'total_vendas': float(total_vendedor_periodo),
                'margem_total': float(margem_vendedor_periodo),
                'margem_percentual': round(float(margem_percentual_periodo), 2)
            })

    queryset = queryset.annotate(
        mes=TruncMonth('data_emissao')
    ).values('mes').annotate(
        total_vendas=Sum('valor_contabil'),
        margem_percentual=Avg(subquery_margem_percentual)*100,
        margem_total=Sum(subquery_margem),
        custo_total=Sum(subquery_custo)
    ).order_by('mes')
    
    labels = [item['mes'].strftime('%b/%Y') for item in queryset]
    
    total_vendas = [float(item['total_vendas']) for item in queryset]
    margem_percentual = [float(item['margem_percentual']) for item in queryset]
    margem_total = [float(item['margem_total']) for item in queryset]
    custo_total = [float(item['custo_total']) for item in queryset] 
        
    return JsonResponse({
        # 'margem': queryset['margem_total'],
        'labels': labels,
        'total_vendas': total_vendas,
        'margem_percentual': margem_percentual,
        'margem_total': margem_total,
        'custo_total': custo_total,
        'estatisticas_justificativas': estatisticas_justificativas,
        'estatisticas_vendedores': estatisticas_vendedores,
        'estatisticas_vendedores_periodo': estatisticas_vendedores_periodo
    })
    

# class OPListView(LoginRequiredMixin, FilterView, ListView):
#     model = OP
#        = ['-ord_producao', '-sequencial']
#     context_object_name = 'op_list'
#     template_name = 'notas/op_list.html'
#     # filterset_class = SeuFiltroDeOPs # Descomente se for usar o FilterView real

#     def get_queryset(self):
#         # 1. Recupera o queryset original com a ordenação definida na classe
#         qs = super().get_queryset()
        
#         # 2. Pega a 'chave' passada na URL
#         lote = self.kwargs.get('lote')
        
#         # 3. Busca a Nota e monta o prefixo
#         # nota = get_object_or_404(Nota, chave=chave)
#         # prefixo_busca = f"{nota.filial}{nota.nota}{nota.item}{nota.recno}"
        
#         # 4. Retorna o queryset filtrado
#         return qs.filter(id_op__startswith=lote)

#     def get_context_data(self, **kwargs):
#         # Adiciona o objeto 'nota' ao contexto para você usar no seu HTML (ex: {{ nota.lote }})
#         context = super().get_context_data(**kwargs)
#         chave = self.kwargs.get('chave')
#         context['nota'] = get_object_or_404(Nota, chave=chave)
#         return context

@login_required
# @cache_page(60 * 15)
def op_list_view(request, lote, cod_produto):
    with open(f'{settings.BASE_DIR}/notas/management/commands/querys/buscaOPs.sql', 'r') as f:
        query_ops = f.read()
        
    with pool.connection() as conn:
        with conn.cursor() as cursor:
            # Passe 'lote' como tupla/lista; passar string direto pode ser
            # tratada como sequência de caracteres pela API do DB.
            cursor.execute(query_ops, (lote,))
            colunas = [coluna[0] for coluna in cursor.description]
            linhas_op = []
            for linha in cursor.fetchall():
                d = dict(zip(colunas, linha))
                for k, v in d.items():
                    if isinstance(v, str):
                        d[k] = v.strip()
                linhas_op.append(d)

    
    if linhas_op:
        existing_ops = {o.id_op: o for o in OP.objects.prefetch_related('custo2_op_set').filter(lote=lote)}
        ops_to_create = []
        ops_to_update = []
            
        update_fields = [c for c in colunas if c != 'id_op']
            
        for op in linhas_op:
            lote_str = str(op.get('lote') or '').strip()
            ord_str = str(op.get('ord_producao') or '').strip()
            seq_str = str(op.get('sequencial') or '').strip()
            tm_str = str(op.get('tp_movimento') or '').strip()
                
            id_op_unico = f"{lote_str}_{ord_str}_{seq_str}_{tm_str}"
            op['id_op'] = id_op_unico
                
            op_data = {k: v for k, v in op.items() if k in update_fields}
                
            for decimal_field in ['quantidade', 'quant_2', 'custo', 'custo_2']:
                if op_data.get(decimal_field) is None:
                    op_data[decimal_field] = 0
                    op[decimal_field] = 0

            if id_op_unico in existing_ops:
                obj = existing_ops[id_op_unico]
                has_custo2_history = obj.pk is not None and len(obj.custo2_op_set.all()) > 0
                if has_custo2_history:
                    op['custo_2'] = obj.custo_2
                    op_data.pop('custo_2', None)

                for k, v in op_data.items():
                    setattr(obj, k, v)
                obj._update_fields = [f for f in update_fields if f in op_data]
                ops_to_update.append(obj)
            else:
                op_data['id_op'] = id_op_unico
                ops_to_create.append(OP(**op_data))
                existing_ops[id_op_unico] = ops_to_create[-1]

        if ops_to_create:
            OP.objects.bulk_create(ops_to_create, batch_size=500)

        if ops_to_update:
            from collections import defaultdict
            grupos = defaultdict(list)
            for obj in ops_to_update:
                campos = tuple(getattr(obj, '_update_fields', update_fields))
                grupos[campos].append(obj)
            for campos, grupo in grupos.items():
                OP.objects.bulk_update(grupo, list(campos), batch_size=500)
    
    for op in linhas_op:
        if op.get('custo') is not None:
            op['custo_raw'] = f"{float(op['custo']):.2f}"
            op['custo'] = locale.currency(op['custo'], grouping=True)
        if op.get('custo_2') is not None:
            # Retrieve latest historical custo_2 if exists
            try:
                op_obj = existing_ops.get(op['id_op'])
                if op_obj and op_obj.pk is not None:
                    custos2 = list(op_obj.custo2_op_set.all())
                    if custos2:
                        latest_custo2 = sorted(custos2, key=lambda x: x.id, reverse=True)[0]
                        op['custo_2'] = latest_custo2.valor
            except Exception:
                pass
            op['custo_2_raw'] = f"{float(op['custo_2']):.2f}"
            op['custo_2'] = locale.currency(op['custo_2'], grouping=True)
        if op.get('quant_2') is not None:
            op['quant_2_raw'] = f"{float(op['quant_2']):.2f}"

        if op.get('c_contabil') is None:
            op['c_contabil'] = '-'
            op['descricao_da_conta'] = '-'
        if op.get('centro_custo') is None:
            op['centro_custo'] = '-'
            op['desc_centro_de_custo'] = '-'
                
    arvore_dict = construir_estado_arvore_producao(linhas_op, cod_produto=cod_produto)
    arvore_json = json.dumps(arvore_dict, default=str)
    
    return render(request, 'notas/op_list.html', {'linhas_op': linhas_op, 'lote': lote, 'cod_produto': cod_produto, 'arvore_json': arvore_json})

class JustificativaListView(LoginRequiredMixin, FilterView, ListView):
    login_url = 'login/'
    context_object_name = 'justificativas'
    template_name = 'notas/justificativas.html'

@login_required
def justificativa_admin_view(request):
    if not request.user.is_superuser:
        return render(request, '403.html', status=403) # Caso haja página de erro
    
    selected = request.GET.get('data_emissao_month', '')
    
# Se não veio valor no GET, tenta preencher com mês atual (se houver dados),
# caso contrário preenche com o mês mais recente disponível.
    if not selected:
        hoje = datetime.datetime.today()
        atual = f"{hoje.year}-{hoje.month:02d}"
        selected = atual
        
    justificativas = Justificativa.objects.all().order_by('-ativo', '-data_cadastro')
    
    logs_justificativas = Nf_Has_Justificativa.objects.all().select_related('usuario', 'nf', 'justificativa').order_by('-data_cadastro')[:200]
    logs_comentarios = Log_Comentario.objects.all().select_related('usuario', 'nf').order_by('-data_cadastro')[:200]
    
    meses_bloqueados = MesBloqueado.objects.all().order_by('-mes_ano')
    logs_bloqueio = LogBloqueioMes.objects.select_related('usuario').order_by('-data_cadastro')[:200]
    meses_qs = Nota.objects.annotate(
        year=ExtractYear('data_emissao'),
        month=ExtractMonth('data_emissao')
    ).values('year', 'month').distinct().order_by('-year', '-month')

    status_por_mes = {
        item.mes_ano: item for item in meses_bloqueados
    }
    meses_gerenciaveis = []

    for item in meses_qs:
        ano = int(item['year'])
        mes = int(item['month'])
        mes_ano = f"{ano}-{mes:02d}"
        status = status_por_mes.get(mes_ano)
        meses_gerenciaveis.append({
            'mes_ano': mes_ano,
            'label': f"{mes:02d}/{ano}",
            'bloqueado': status.bloqueado if status else False,
            'data_atualizacao': status.data_atualizacao if status else None,
        })
        
    return render(request, 'notas/admin.html', {
        'justificativas': justificativas, 
        'logs_justificativas': logs_justificativas,
        'logs_comentarios': logs_comentarios,
        'meses_bloqueados': meses_bloqueados,
        'meses_gerenciaveis': meses_gerenciaveis,
        'logs_bloqueio': logs_bloqueio,
        'selected_month': selected,
        'nota_grid_columns': NOTA_GRID_COLUMNS,
        'nota_grid_visible_columns': get_nota_grid_visible_columns(),
    })

@login_required
@require_POST
def salvar_preferencia_colunas_nota(request):
    try:
        data = json.loads(request.body)
        colunas = data.get('colunas_visiveis', [])

        if not isinstance(colunas, list):
            return JsonResponse({'success': False, 'error': 'Formato de colunas inválido.'}, status=400)

        colunas_validas = []
        for coluna in colunas:
            if coluna in NOTA_GRID_COLUMN_KEYS and coluna not in colunas_validas:
                colunas_validas.append(coluna)

        if not colunas_validas:
            return JsonResponse({'success': False, 'error': 'Selecione ao menos uma coluna.'}, status=400)

        preferencia = PreferenciaGlobalColunasNota.objects.first()
        if not preferencia:
            preferencia = PreferenciaGlobalColunasNota()
            
        preferencia.colunas_visiveis = colunas_validas
        preferencia.save(update_fields=['colunas_visiveis', 'data_atualizacao'] if preferencia.pk else None)

        return JsonResponse({'success': True, 'colunas_visiveis': colunas_validas})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def justificativa_save(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Acesso negado'}, status=403)
    
    data = json.loads(request.body)
    j_id = data.get('id')
    texto = data.get('texto')
    
    if not texto:
        return JsonResponse({'success': False, 'error': 'Texto obrigatório'}, status=400)
        
    try:
        if j_id:
            justificativa = Justificativa.objects.get(id=j_id)
            justificativa.texto = texto
            justificativa.usuario = CustomUser.objects.get(id=request.user.id)
            justificativa.save()
        else:
            Justificativa.objects.create(texto=texto, usuario=request.user)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def justificativa_toggle_status(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Acesso negado'}, status=403)
        
    data = json.loads(request.body)
    j_id = data.get('id')
    
    try:
        justificativa = Justificativa.objects.get(id=j_id)
        justificativa.ativo = not justificativa.ativo
        justificativa.save()
        return JsonResponse({'success': True, 'ativo': justificativa.ativo})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
   

@login_required
@require_POST
def toggle_bloqueio_mes_api(request):
    """Nova View para gerenciar o bloqueio de meses"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Acesso negado'}, status=403)
        
    try:
        data = json.loads(request.body)
        mes_ano = data.get('mes_ano') # Formato: YYYY-MM
        
        if not mes_ano:
            return JsonResponse({'success': False, 'error': 'Mês não fornecido'}, status=400)
            
        mes_status, created = MesBloqueado.objects.get_or_create(mes_ano=mes_ano)
        mes_status.bloqueado = not mes_status.bloqueado
        mes_status.save()
        
        acao_str = 'BLOQUEADO' if mes_status.bloqueado else 'DESBLOQUEADO'
        
        LogBloqueioMes.objects.create(
            mes_ano=mes_ano,
            acao=acao_str,
            usuario=request.user
        )
        
        return JsonResponse({'success': True, 'bloqueado': mes_status.bloqueado, 'mes': mes_ano})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
 

@login_required
def exportar_excel(request):
    import pandas as pd
    import io
    from xlsxwriter.utility import xl_col_to_name

    # Obtém o queryset base com as pre-anotações e filial filter
    notas_view = NotasListView()
    notas_view.request = request
    base_qs = notas_view.get_queryset()

    # Aplica os filtros da URL
    f = NotaFilter(request.GET, queryset=base_qs)
    queryset = f.qs

    columns = [
        'chave', 'filial', 'nome_filial', 'nota', 'item', 'no_pedido', 'vendedor',
        'data_emissao', 'lote', 'cfop', 'cfop_descri', 'atualiza_estoque', 'gera_duplicata',
        'cod_produto', 'produto', 'tipo_produto', 'desc_tipo_produto', 'armazem',
        'cod_cliente', 'loja', 'cliente', 'grp_amar_ctb', 'classificacao_produto',
        'estado_destino', 'quantidade', 'tabela_preco', 'preco_tabela',
        'valor_contabil', 'custo_mais_recente', 'valor_unitario', 'valor_ipi', 'valor_imp5',
        'valor_imp6', 'valor_icms_difal', 'valor_icms', 'base_icms','aliq_icms',
        'margem_bruta', 'margem_percentual'
    ]

    dados = list(queryset.values(*columns))
    df = pd.DataFrame(dados)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        sheet_name = "NFs"

        if df.empty:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            if 'data_emissao' in df.columns:
                df['data_emissao'] = pd.to_datetime(df['data_emissao']).dt.date
            
            cols_to_float = [
                'quantidade', 'preco_tabela', 'valor_contabil', 'custo_mais_recente', 
                'valor_unitario', 'valor_ipi', 'valor_imp5', 'valor_imp6', 
                'valor_icms_difal', 'valor_icms',  'base_icms', 'aliq_icms', 'margem_bruta', 'margem_percentual'
            ]
            for col in cols_to_float:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

            if 'aliq_icms' in df.columns:
                df['aliq_icms'] = df['aliq_icms'] / 100.0

            if 'margem_percentual' in df.columns:
                df['margem_percentual'] = df['margem_percentual'] / 100.0

            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            fmt_header = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1})
            fmt_currency = workbook.add_format({'num_format': 'R$ #,##0.00'}) 
            fmt_percent = workbook.add_format({'num_format': '0.00%'})
            fmt_date = workbook.add_format({'num_format': 'dd/mm/yyyy'})
            
            cor_ruim = '#FFC7CE'
            texto_ruim = '#9C0006'
            cor_alta = '#FFEB9C'
            texto_alta = '#9C6500'

            formats = {
                'red': {
                    'geral': workbook.add_format({'bg_color': cor_ruim, 'font_color': texto_ruim}),
                    'money': workbook.add_format({'bg_color': cor_ruim, 'font_color': texto_ruim, 'num_format': 'R$ #,##0.00'}),
                    'percent': workbook.add_format({'bg_color': cor_ruim, 'font_color': texto_ruim, 'num_format': '0.00%'}),
                    'date': workbook.add_format({'bg_color': cor_ruim, 'font_color': texto_ruim, 'num_format': 'dd/mm/yyyy'})
                },
                'yellow': {
                    'geral': workbook.add_format({'bg_color': cor_alta, 'font_color': texto_alta}),
                    'money': workbook.add_format({'bg_color': cor_alta, 'font_color': texto_alta, 'num_format': 'R$ #,##0.00'}),
                    'percent': workbook.add_format({'bg_color': cor_alta, 'font_color': texto_alta, 'num_format': '0.00%'}),
                    'date': workbook.add_format({'bg_color': cor_alta, 'font_color': texto_alta, 'num_format': 'dd/mm/yyyy'})
                },
                'blank': {
                    'geral': workbook.add_format({}),
                    'money': fmt_currency,
                    'percent': fmt_percent,
                    'date': fmt_date
                }
            }

            for idx, col in enumerate(df.columns):
                series = df[col]
                max_len = max((series.astype(str).map(len).max(), len(str(col)))) + 2
                worksheet.set_column(idx, idx, max_len)
                worksheet.write(0, idx, col, fmt_header)

            set_parceiros = set()
            try:
                with pool_mysql.connection() as con:
                    with con.cursor() as cursor:
                        query = """SELECT cod_cliente, loja_cliente FROM analise_margem.cliente_parceiro;"""
                        cursor.execute(query)
                        clientes_parceiros_raw = cursor.fetchall()
                set_parceiros = {(str(c[0]).strip(), str(c[1]).strip()) for c in clientes_parceiros_raw}
            except Exception:
                pass

            col_letters = {col: xl_col_to_name(idx) for idx, col in enumerate(df.columns)}
            
            col_cfop = col_letters.get('cfop', 'A')
            col_valor_contabil = col_letters.get('valor_contabil', 'A')
            col_custo = col_letters.get('custo_mais_recente', 'A')
            col_ipi = col_letters.get('valor_ipi', 'A')
            col_imp5 = col_letters.get('valor_imp5', 'A')
            col_imp6 = col_letters.get('valor_imp6', 'A')
            col_icms_difal = col_letters.get('valor_icms_difal', 'A')
            col_icms = col_letters.get('valor_icms', 'A')
            col_base_icms = col_letters.get('base_icms', 'A')
            
            idx_margem_bruta = df.columns.get_loc('margem_bruta') if 'margem_bruta' in df.columns else -1
            idx_margem_percentual = df.columns.get_loc('margem_percentual') if 'margem_percentual' in df.columns else -1

            for row_idx, row in enumerate(df.itertuples(index=False), start=1):
                margem = getattr(row, 'margem_percentual', 0)
                cod_cliente = str(getattr(row, 'cod_cliente', '')).strip()
                loja = str(getattr(row, 'loja', '')).strip()
                parceiro = (cod_cliente, loja) in set_parceiros
                cod_produto = getattr(row, 'cod_produto', None)
                meta_minima = 0.17 if parceiro else 0.27
                
                tipo_destaque = 'blank'
                if cod_produto not in ('B0010046','E000H2P8'):
                    if margem < meta_minima:
                        tipo_destaque = 'red'
                    elif margem > 0.50:
                        tipo_destaque = 'yellow'
                        
                dict_formatos = formats[tipo_destaque]
                
                for col_idx, col_name in enumerate(df.columns):
                    valor_celula = row[col_idx]
                    col_lower = col_name.lower()
                    formato_final = dict_formatos['geral']
                    
                    if any(x in col_lower for x in ['margem_percentual', 'margem_bruta_percentual', 'aliq_icms']):
                        formato_final = dict_formatos['percent']
                    elif any(x in col_lower for x in ['valor_', 'custo', 'vlr_', 'margem_bruta', 'preco']):
                        formato_final = dict_formatos['money']
                    elif any(x in col_lower for x in ['data_emissao']):
                        formato_final = dict_formatos['date']
                        
                    worksheet.write(row_idx, col_idx, valor_celula, formato_final)
                    
                if idx_margem_bruta != -1:
                    formula_margem = (
                        f'={col_valor_contabil}{row_idx+1}-{col_custo}{row_idx+1}-{col_ipi}{row_idx+1}-'
                        f'{col_imp5}{row_idx+1}-{col_imp6}{row_idx+1}-{col_icms_difal}{row_idx+1}-'
                        f'IF(OR({col_cfop}{row_idx+1}="5101", {col_cfop}{row_idx+1}="6101", '
                        f'{col_cfop}{row_idx+1}="5116", {col_cfop}{row_idx+1}="6116", '
                        f'{col_cfop}{row_idx+1}="6107"), {col_icms}{row_idx+1}*0.047, {col_icms}{row_idx+1})'
                    )
                    worksheet.write_formula(row_idx, idx_margem_bruta, formula_margem, dict_formatos['money'])
                
                if idx_margem_percentual != -1 and idx_margem_bruta != -1:
                    formula_margem_pct = f'=IF({col_valor_contabil}{row_idx+1}=0, 0, {xl_col_to_name(idx_margem_bruta)}{row_idx+1}/{col_valor_contabil}{row_idx+1})'
                    worksheet.write_formula(row_idx, idx_margem_percentual, formula_margem_pct, dict_formatos['percent'])

            worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    hoje = datetime.datetime.today()
    atual = f"{hoje.year}-{hoje.month:02d}"
    response['Content-Disposition'] = f'attachment; filename="NFs-{atual}.xlsx"'
    return response


@login_required
def exportar_estatisticas_excel(request):
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

# 1. Obtém os dados chamando a API localmente
    resp = dados_vendas_api(request)
    if resp.status_code != 200:
        return HttpResponse("Erro ao gerar dados", status=500)
    
    dados_json = json.loads(resp.content)
    
    wb = Workbook()
    
# Aba 1: Estatísticas Justificativas
    ws1 = wb.active
    ws1.title = "Estatísticas Justificativas"
    
    cols1 = ['Justificativa', 'Qtd Notas', 'Representatividade Vendas', 'Abaixo Margem Percentual']
    ws1.append(cols1)
    
    for item in dados_json.get('estatisticas_justificativas', []):
        ws1.append([
            item['justificativa'],
            item['contagem'],
            item['representatividade_vendas'] / 100.0,
            item['percentual_abaixo_margem'] / 100.0
        ])
        
# Aba 2: Estatísticas Período
    ws2 = wb.create_sheet(title="Estatísticas Por Período")
    cols2 = ['Mês', 'Faturamento (valor_)', 'Custo Total', 'Margem Bruta', 'Margem Percentual']
    ws2.append(['Mês', 'Faturamento', 'Custo Total', 'Margem Total', 'Margem Percentual'])
    
    labels = dados_json.get('labels', [])
    vendas = dados_json.get('total_vendas', [])
    custos = dados_json.get('custo_total', [])
    margens = dados_json.get('margem_total', [])
    margens_pct = dados_json.get('margem_percentual', [])
    
    for i in range(len(labels)):
        ws2.append([
            labels[i],
            vendas[i],
            custos[i],
            margens[i],
            (margens_pct[i] / 100.0) if margens_pct[i] is not None else 0
        ])

# Aba 3: Estatísticas Vendedores
    ws3 = wb.create_sheet(title="Estatísticas Por Vendedor")
    cols3 = ['Vendedor', 'Qtd Notas', 'Faturamento', 'Custo Total', 'Margem Total', 'Margem Percentual']
    ws3.append(cols3)

    for item in dados_json.get('estatisticas_vendedores', []):
        ws3.append([
            item['vendedor'],
            item['qtd_notas'],
            item['total_vendas'],
            item['custo_total'],
            item['margem_total'],
            item['margem_percentual'] / 100.0
        ])

# Aba 4: Evolução Vendedores
    ws4 = wb.create_sheet(title="Evolução Por Vendedor")
    cols4 = ['Mês', 'Vendedor', 'Faturamento', 'Margem Total', 'Margem Percentual']
    ws4.append(cols4)

    for item in dados_json.get('estatisticas_vendedores_periodo', []):
        ws4.append([
            item['mes'],
            item['vendedor'],
            item['total_vendas'],
            item['margem_total'],
            item['margem_percentual'] / 100.0
        ])
    
# Aplicar formato
    fmt_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fmt_header_font = Font(bold=True)
    fmt_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for ws, cols in [(ws1, cols1), (ws2, cols2), (ws3, cols3), (ws4, cols4)]:
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.font = fmt_header_font
            cell.fill = fmt_header_fill
            cell.border = fmt_border

        for col_idx, col_name in enumerate(cols, 1):
            col_name_lower = str(col_name).lower()
            num_format = None
            if any(x in col_name_lower for x in ['valor_', 'custo', 'vlr_', 'margem_bruta', 'margem total', 'faturamento']):
                num_format = 'R$ #,##0.00'
            elif any(x in col_name_lower for x in ['percent', 'aliquota', 'diff_percentual', 'representatividade', 'abaixo margem']):
                num_format = '0.00%'

            max_len = len(str(ws.cell(row=1, column=col_idx).value))
            
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if num_format:
                    cell.number_format = num_format
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
            
        last_col_letter = get_column_letter(len(cols))
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estatisticas.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_op_excel(request, lote):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
    
    with open(f'{settings.BASE_DIR}/notas/management/commands/querys/buscaOPs.sql', 'r') as file:
        query_ops = file.read()
        
    with pool.connection() as conn:
        with conn.cursor() as cursor:
            # Passe 'lote' como tupla/lista; passar string direto pode ser
            # tratada como sequência de caracteres pela API do DB.
            cursor.execute(query_ops, (lote,))
            colunas = [coluna[0] for coluna in cursor.description]
            linhas_op = [dict(zip(colunas, linha)) for linha in cursor.fetchall()]

    if linhas_op:
        existing_ops = {o.id_op: o for o in OP.objects.prefetch_related('custo2_op_set').filter(lote=lote)}
        for op in linhas_op:
            lote_str = str(op.get('lote') or '').strip()
            ord_str = str(op.get('ord_producao') or '').strip()
            seq_str = str(op.get('sequencial') or '').strip()
            tm_str = str(op.get('tp_movimento') or '').strip()
            id_op_unico = f"{lote_str}_{ord_str}_{seq_str}_{tm_str}"
            
            if id_op_unico in existing_ops and existing_ops[id_op_unico].custo2_op_set.exists():
                op['custo_2'] = existing_ops[id_op_unico].custo_2

    wb = Workbook()
    ws = wb.active
    ws.title = f"OPs Lote {lote}"
    
    cols_display = [
        ('filial', 'Filial'),
        ('produto', 'Produto'),
        ('descr_prod', 'Descr. Produto'),
        ('armazem', 'Armazém'),
        ('tp_movimento', 'Tp. Mov.'),
        ('descricao_tm', 'Desc. TM'),
        ('unidade', 'Unidade'),
        ('quantidade', 'Quantidade'),
        ('quant_2', 'Quantidade 2'),
        ('custo', 'Custo'),
        ('custo_2', 'Custo 2'),
        ('ord_producao', 'Ord. Produção'),
        ('lote', 'Lote'),
        ('os_ass_tecn', 'OS Ass. Tecn.'),
        ('grupo', 'Grupo'),
        ('descricao_grupo', 'Desc. Grupo'),
        ('tipo_re_de', 'Tipo RE/DE'),
        ('ext_texto', 'Ext. Texto'),
        ('documento', 'Documento'),
        ('dt_emissao', 'Dt. Emissão'),
        ('c_contabil', 'C. Contábil'),
        ('descricao_da_conta', 'Desc. Conta'),
        ('centro_custo', 'Centro Custo'),
        ('desc_centro_de_custo', 'Desc. C.Custo'),
        ('parc_total', 'Parc./Total'),
        ('estornado', 'Estornado'),
        ('sequencial', 'Sequencial'),
        ('tipo', 'Tipo'),
        ('usuario', 'Usuário'),
        ('nr_s_a', 'Nr. S.A.'),
        ('item_s_a', 'Item S.A.'),
    ]
    
    ws.append([c[1] for c in cols_display])
    
    for op in linhas_op:
        row_data = []
        for key, name in cols_display:
            val = op.get(key)
            if key in ['c_contabil', 'centro_custo'] and val is None:
                val = '-'
            elif key in ['descricao_da_conta', 'desc_centro_de_custo'] and val is None:
                val = '-'
            row_data.append(val)
        ws.append(row_data)
        
    fmt_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fmt_header_font = Font(bold=True)
    fmt_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.font = fmt_header_font
        cell.fill = fmt_header_fill
        cell.border = fmt_border
        
    col_names = [c[1] for c in cols_display]
    for col_idx, col_name in enumerate(col_names, 1):
        col_name_lower = str(col_name).lower()
        num_format = None
        
        if any(x in col_name_lower for x in ['valor_', 'custo', 'vlr_', 'margem_bruta']):
            num_format = 'R$ #,##0.00'
        elif any(x in col_name_lower for x in ['percent', 'aliquota', 'diff_percentual']):
            num_format = '0.00%'

        max_len = len(str(col_name))
        
        last_row = len(linhas_op) + 1
        for row_idx in range(2, last_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if num_format:
                cell.number_format = num_format
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    if 'Tp. Mov.' in col_names:
        col_idx = col_names.index('Tp. Mov.') + 1
        col_letter = get_column_letter(col_idx)
        
        fmt_destaque = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        last_row = len(linhas_op) + 1
        last_col_letter = get_column_letter(len(col_names))
        
        formula = [f'=${col_letter}2="010"']
        rule = FormulaRule(formula=formula, stopIfTrue=True, fill=fmt_destaque)
        ws.conditional_formatting.add(f"A2:{last_col_letter}{last_row}", rule)

    last_col_letter = get_column_letter(len(col_names))
    last_row = len(linhas_op) + 1
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="OPs_lote_{lote}.xlsx"'
    wb.save(response)
    return response

FlatRow = dict[str, Any]


class FolhaIndexada(TypedDict):
    id_nivel_1: str
    quantidade_acumulada: float
    custo_atual_folha: float


class EstadoNivel1(TypedDict):
    custo_medio_base: float
    custo_medio: float


def _to_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ''):
        return default
    if isinstance(value, str):
        value = value.replace('R$', '').strip()
        if ',' in value:
            value = value.replace('.', '').replace(',', '.')
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _custo_medio_erp(row: FlatRow) -> float:
    return _to_float(row.get('custo_2_raw', row.get('custo_2', 0)))


def _quantidade_item(row: FlatRow) -> float:
    return _to_float(row.get('quant_2_raw', row.get('quant_2', row.get('quantidade', 1))), 1.0)


def construir_estado_arvore_producao(
    flat_data: list[FlatRow],
    cod_produto: str | None = None,
) -> dict[str, Any]:
    ops: dict[str, dict[str, Any]] = {}
    produtos_consumidos: set[str] = set()
    op_por_produto: dict[str, str] = {}

    for row in flat_data:
        op_id = row['ord_producao']
        op_data = ops.setdefault(op_id, {'pai': None, 'filhos': []})

        if row['tp_movimento'] == '010':
            op_data['pai'] = row
            op_por_produto.setdefault(row['produto'], op_id)
        else:
            op_data['filhos'].append(row)
            produtos_consumidos.add(row['produto'])

    root_op = None
    for op_id, dados in ops.items():
        pai = dados['pai']
        if pai and cod_produto and pai['produto'] == cod_produto:
            root_op = op_id
            break
        if pai and not cod_produto and pai['produto'] not in produtos_consumidos:
            root_op = op_id
            break

    folhas_indexadas: dict[str, list[FolhaIndexada]] = {}
    estado_nivel_1: dict[str, EstadoNivel1] = {}

    def registrar_nivel_1(id_nivel_1: str, custo_base: float) -> None:
        estado = estado_nivel_1.setdefault(
            id_nivel_1,
            {'custo_medio_base': 0.0, 'custo_medio': 0.0},
        )
        estado['custo_medio_base'] += custo_base
        estado['custo_medio'] += custo_base

    def registrar_folha(row: FlatRow, id_nivel_1: str, quantidade_acumulada: float) -> None:
        folhas_indexadas.setdefault(row['produto'], []).append({
            'id_nivel_1': id_nivel_1,
            'quantidade_acumulada': quantidade_acumulada,
            'custo_atual_folha': _custo_medio_erp(row),
        })

    def montar_componente(
        row: FlatRow,
        nivel: int,
        id_nivel_1: str,
        quantidade_acumulada: float,
    ) -> FlatRow:
        produto = row['produto']
        op_do_filho = op_por_produto.get(produto)
        no_atual = row.copy()
        no_atual['nivel'] = nivel
        no_atual['quantidade_acumulada'] = quantidade_acumulada
        no_atual['filhos'] = []

        if nivel == 1:
            # Baseline absoluto do L1: sempre o custo médio da linha flat do ERP.
            custo_base = _custo_medio_erp(row)
            registrar_nivel_1(id_nivel_1, custo_base)
            no_atual['custo_medio_base'] = custo_base
            no_atual['custo_medio'] = custo_base
            no_atual['custo_calculado'] = custo_base
        else:
            no_atual['custo_calculado'] = _custo_medio_erp(row)

        if op_do_filho and op_do_filho in ops:
            for filho in ops[op_do_filho]['filhos']:
                quantidade_filho = quantidade_acumulada * _quantidade_item(filho)
                no_atual['filhos'].append(
                    montar_componente(
                        filho,
                        nivel + 1,
                        id_nivel_1,
                        quantidade_filho,
                    )
                )
            no_atual['is_leaf'] = False
        else:
            no_atual['is_leaf'] = True
            registrar_folha(row, id_nivel_1, quantidade_acumulada)

        return no_atual

    if not root_op or not ops[root_op]['pai']:
        return {
            'arvore': {},
            'folhas_indexadas': folhas_indexadas,
            'estado_nivel_1': estado_nivel_1,
            'custo_raiz': 0.0,
        }

    arvore = ops[root_op]['pai'].copy()
    arvore['nivel'] = 0
    arvore['quantidade_acumulada'] = 1.0
    arvore['filhos'] = []
    arvore['is_leaf'] = False

    for filho in ops[root_op]['filhos']:
        arvore['filhos'].append(
            montar_componente(
                filho,
                nivel=1,
                id_nivel_1=filho['produto'],
                quantidade_acumulada=1.0,
            )
        )

    custo_raiz = sum(estado['custo_medio'] for estado in estado_nivel_1.values())
    arvore['custo_medio'] = custo_raiz
    arvore['custo_calculado'] = custo_raiz

    return {
        'arvore': arvore,
        'folhas_indexadas': folhas_indexadas,
        'estado_nivel_1': estado_nivel_1,
        'custo_raiz': custo_raiz,
    }


def construir_arvore_producao(flat_data: list[FlatRow], cod_produto: str | None = None) -> FlatRow:
    return construir_estado_arvore_producao(flat_data, cod_produto=cod_produto)['arvore']


def simular_variacao_custo(
    folhas_indexadas: dict[str, list[FolhaIndexada]],
    estado_nivel_1: dict[str, EstadoNivel1],
    custo_raiz: float,
    id_folha_editada: str,
    novo_custo_folha: float,
) -> float:
    caminhos = folhas_indexadas.get(id_folha_editada, [])

    for caminho in caminhos:
        delta_unitario = novo_custo_folha - caminho['custo_atual_folha']
        delta_total = delta_unitario * caminho['quantidade_acumulada']

        estado_nivel_1[caminho['id_nivel_1']]['custo_medio'] += delta_total
        custo_raiz += delta_total
        caminho['custo_atual_folha'] = novo_custo_folha

    return custo_raiz
